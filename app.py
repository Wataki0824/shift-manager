#!/usr/bin/env python3

"""
勤務表自動生成 Webアプリ
Author: Taiki Watanabe
"""

import os
import io
import random
from pathlib import Path
from functools import wraps

import pandas as pd
from flask import Flask, render_template, request, send_file, jsonify, session, redirect, url_for
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# パスワード設定（環境変数から取得，デフォルトは 'kitakyushu'）
APP_PASSWORD = os.environ.get('APP_PASSWORD', 'kitakyusyu')


def login_required(f):
    """ログイン必須デコレータ"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('authenticated'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


# ===== 勤務表生成ロジック（generate_shift.pyから移植） =====

def is_weekend(day_of_week: str) -> bool:
    """土日かどうか判定"""
    return day_of_week in ['土', '日']


def is_off(value: str) -> bool:
    """休みかどうか判定（公，年，または空でない休み表記）"""
    if pd.isna(value) or value == '':
        return False
    return value in ['公', '年']


def get_off_value(value: str) -> str:
    """休みの値を取得（公 or 年）"""
    if pd.isna(value) or value == '':
        return ''
    if value in ['公', '年']:
        return value
    return ''


def select_y_candidate(candidates: list, y_counts: dict, prev_day_off: dict) -> str:
    """呼び出しを割り当てる候補を選ぶ"""
    if not candidates:
        raise ValueError("候補者がいません")
    
    if len(candidates) == 1:
        return candidates[0]
    
    not_after_off = [c for c in candidates if not prev_day_off.get(c, False)]
    
    if not_after_off:
        target_list = not_after_off
    else:
        target_list = candidates
    
    min_y = min(y_counts[c] for c in target_list)
    min_y_candidates = [c for c in target_list if y_counts[c] == min_y]
    
    return min_y_candidates[0]


def fill_missing_public_holidays(df: pd.DataFrame, staff_columns: list, target_holidays: int = 8, headcount: dict = None) -> pd.DataFrame:
    """公休を自動的に埋める"""
    # 指定日数になるまでランダムに '公' を入れる
    # ただし以下のルールを守る
    # 1. 土日のどちらかは必ず出勤（土日両方休みはNG、土日連勤も可能な限り避けるが、
    #    「土日は最低1人出勤」の制約があるため、全員公休はNG）
    # 2. 7連勤以上はNG -> 6連勤まで
    
    # 修正: 平日も最低2人出勤 -> 3人中1人休みまで
    
    result_df = df.copy()
    holiday_counts = {staff: result_df[staff].eq('公').sum() + result_df[staff].eq('年').sum() for staff in staff_columns}
    
    # --- Phase -1: 人数指定 (headcount) の適用 ---
    # 指定がある日は、その人数になるように公休を入れる（または入れない）
    if headcount:
        for idx, row in result_df.iterrows():
            day_str = str(row['日付']) # 文字列として比較
            target_work_count = int(headcount.get(day_str, 0))
            
            if target_work_count > 0:
                # 現在の休み状況確認
                current_off = [s for s in staff_columns if get_off_value(row[s])]
                current_off_count = len(current_off)
                
                # 必要な休み人数 = 全員 - 目標出勤人数
                # 例: 3人 - 3人出勤 = 0人休み
                # 例: 3人 - 1人出勤 = 2人休み
                needed_off_count = len(staff_columns) - target_work_count
                
                # 負の値防止
                needed_off_count = max(0, needed_off_count)
                
                if needed_off_count > current_off_count:
                    # 休みを増やす
                    diff = needed_off_count - current_off_count
                    # 公休候補: 既に休みでなく、かつ固定シフト(9, 10)でもない人
                    candidates = []
                    for s in staff_columns:
                        val = result_df.at[idx, s]
                        if not get_off_value(val) and val not in ['9', '10']:
                            candidates.append(s)
                            
                    if len(candidates) >= diff:
                        to_be_off = random.sample(candidates, diff)
                        for staff in to_be_off:
                            result_df.at[idx, staff] = '公'
                            holiday_counts[staff] += 1
                
                # 既に休みが多すぎる場合（例えば3人出勤指定なのに1人公休がいる）
                # ユーザー指定の公休は消さない方針なので、何もしない（出勤人数不足を受け入れる）
                
    # Phase 0: 土日の出勤人数バランス調整
    # ただし、人数指定がある日はスキップする
    weekend_indices = []
    for idx, row in result_df.iterrows():
        day_str = str(row['日付'])
        if headcount and int(headcount.get(day_str, 0)) > 0:
            continue
        if is_weekend(row['曜日']):
            weekend_indices.append(idx)
            
    # ...(中略)...
    
    # Phase 1 & 2 & 3 のループ内でも人数指定日をスキップする必要がある
    # ここでは fill_missing_public_holidays の後半ロジック全体を見直す必要があるため
    # 以下のようにチェックを追加する関数内修正を行う


    # 各スタッフの不足公休数を計算（管理用）
    missing_counts = {}
    for staff in staff_columns:
        current_holidays = result_df[staff].apply(lambda x: 1 if x == '公' else 0).sum()
        target = target_holidays
        missing_counts[staff] = max(0, target - current_holidays)

    # ==========================================
    # Phase 0: 土日出勤の調整（全体最適）
    # ==========================================
    
    weekend_indices = []
    for idx, row in result_df.iterrows():
        # 人数指定がある日は調整対象外（ユーザー指定を優先）
        day_str = str(row['日付'])
        if headcount and int(headcount.get(day_str, 0)) > 0:
            continue
            
        if is_weekend(row['曜日']):
            weekend_indices.append(idx)
    
    for _ in range(100):
        adjusted_any = False
        random.shuffle(weekend_indices)
        
        for idx in weekend_indices:
            working_staff = []
            for s in staff_columns:
                if not is_off(result_df.at[idx, s]):
                    working_staff.append(s)
            
            if len(working_staff) > 1:
                candidates = [s for s in working_staff if missing_counts[s] > 0]
                if not candidates:
                    continue
                
                # 土日休みが少ない人を優先
                weekend_off_counts = {}
                for s in candidates:
                    count = 0
                    for _idx, _row in result_df.iterrows():
                        if is_weekend(_row['曜日']) and is_off(_row[s]):
                            count += 1
                    weekend_off_counts[s] = count
                
                min_off = min(weekend_off_counts.values())
                target_candidates = [s for s in candidates if weekend_off_counts[s] == min_off]
                target_staff = random.choice(target_candidates)
                
                result_df.at[idx, target_staff] = '公'
                missing_counts[target_staff] -= 1
                adjusted_any = True
        
        if not adjusted_any:
            break

    # ==========================================
    # Phase 1 & 2: 連勤解消 & 残り埋め
    # ==========================================
    
    for staff in staff_columns:
        missing = missing_counts[staff]
        if missing <= 0:
            continue
            
        added_count = 0
        limits_to_check = [6, 5] 
        
        for limit in limits_to_check:
            if added_count >= missing:
                break
                
            while added_count < missing:
                # 連勤検知
                consecutive_indices = []
                current_streak = []
                for idx, row in result_df.iterrows():
                    if not is_off(row[staff]):
                        current_streak.append(idx)
                    else:
                        if len(current_streak) > limit:
                            consecutive_indices.append(current_streak.copy())
                        current_streak = []
                if len(current_streak) > limit:
                     consecutive_indices.append(current_streak.copy())
                
                if not consecutive_indices:
                    break
                
                filled_something = False
                for streak in consecutive_indices:
                    if added_count >= missing:
                         break
                    candidates = streak.copy()
                    random.shuffle(candidates)
                    for idx in candidates:
                        # 人数指定がある日はスキップ
                        day_str = str(result_df.at[idx, '日付'])
                        if headcount and int(headcount.get(day_str, 0)) > 0:
                            continue

                        if can_take_off(result_df, idx, staff, staff_columns):
                            result_df.at[idx, staff] = '公'
                            added_count += 1
                            filled_something = True
                            break 
                if not filled_something:
                    break

        # --- Phase 3: まだ足りなければランダムに埋める ---
        if added_count < missing:
            candidate_indices = []
            for idx, row in result_df.iterrows():
                # 人数指定がある日はスキップ
                day_str = str(row['日付'])
                if headcount and int(headcount.get(day_str, 0)) > 0:
                    continue
                    
                val = row[staff]
                # 固定シフト('9', '10')は公休にしない
                if not is_off(val) and val not in ['9', '10']:
                    candidate_indices.append(idx)
            
            random.shuffle(candidate_indices)
            
            for idx in candidate_indices:
                if added_count >= missing:
                    break
                if can_take_off(result_df, idx, staff, staff_columns):
                    result_df.at[idx, staff] = '公'
                    added_count += 1
            
    return result_df


def can_take_off(df: pd.DataFrame, idx: int, staff: str, staff_columns: list) -> bool:
    """指定した日にそのスタッフが休んでも大丈夫かチェック"""
    day_of_week = df.at[idx, '曜日']
    is_wknd = is_weekend(day_of_week)
    
    # 他の出勤スタッフを確認
    working_staff = []
    for s in staff_columns:
        if s == staff:
            continue
        if not is_off(df.at[idx, s]):
            working_staff.append(s)
    
    working_count = len(working_staff)
    
    if is_wknd:
        # 土日は最低1人必要
        return working_count >= 1
    else:
        # 平日は最低2人必要（1人出勤はNG）
        return working_count >= 2



def optimize_shifts(df: pd.DataFrame, staff_columns: list, fixed_mask: dict = None) -> pd.DataFrame:
    """生成後のシフトを微調整して最適化する"""
    # 主に「休み明けのY」を回避するためのスワップを行う
    result_df = df.copy()
    
    for idx, row in result_df.iterrows():
        day_of_week = row['曜日']
        if is_weekend(day_of_week):
            continue # 土日は1人出勤が基本なのでスワップ余地なし
            
        # この日の出勤者とシフト情報を取得
        workers = []
        for staff in staff_columns:
            val = result_df.at[idx, staff]
            if not is_off(val):
                workers.append({'staff': staff, 'val': val})
        
        # 2人以上出勤していないとスワップできない
        if len(workers) < 2:
            continue
            
        # Yがついている人を探す
        y_staff_info = None
        normal_staff_info = None
        
        # 平日で「Y付きの人」と「Yなしの人」のペアを探す
        for w in workers:
            if 'Y' in w['val']:
                y_staff_info = w
            else:
                normal_staff_info = w 
        
        if not y_staff_info or not normal_staff_info:
            continue
            
        y_staff = y_staff_info['staff']
        norm_staff = normal_staff_info['staff']
        
        # 固定シフトが含まれている場合はスワップ不可
        if fixed_mask:
            if fixed_mask.get((idx, y_staff)) or fixed_mask.get((idx, norm_staff)):
                continue

        # 前日の状態を確認
        if idx == 0:
            continue
            
        prev_idx = idx - 1
        y_staff_prev_off = is_off(result_df.at[prev_idx, y_staff])
        norm_staff_prev_off = is_off(result_df.at[prev_idx, norm_staff])
        
        # 「Y担当が前日休み」かつ「通常担当が前日出勤」の場合、入れ替えるべき
        if y_staff_prev_off and not norm_staff_prev_off:
            # スワップ実行
            val_y = y_staff_info['val']
            val_norm = normal_staff_info['val']
            
            result_df.at[idx, y_staff] = val_norm
            result_df.at[idx, norm_staff] = val_y

    return result_df


def assign_shifts(df: pd.DataFrame, staff_columns: list) -> pd.DataFrame:
    """勤務表を自動生成する"""
    result_df = df.copy()

    y_counts = {staff: 0 for staff in staff_columns}
    prev_day_off = {staff: False for staff in staff_columns}
    
    for idx, row in df.iterrows():
        day_of_week = row['曜日']
        weekend = is_weekend(day_of_week)
        
        staff_off = {}
        fixed_shift = {}
        staff_working = []
        
        for staff in staff_columns:
            off_value = get_off_value(row[staff])
            if off_value:
                staff_off[staff] = off_value
            elif row[staff] in ['9', '10']:
                fixed_shift[staff] = row[staff]
                staff_working.append(staff)
                # 固定値を一旦セット（後で上書きされる可能性もあるが基本維持）
                result_df.at[idx, staff] = row[staff]
            else:
                staff_working.append(staff)
        
        # 休み情報を確定
        for staff, off_value in staff_off.items():
            result_df.at[idx, staff] = off_value
        
        if not staff_working:
            for staff in staff_columns:
                prev_day_off[staff] = is_off(row[staff])
            continue
        
        if len(staff_working) == 1:
            only_staff = staff_working[0]
            result_df.at[idx, only_staff] = 'Y9'
            y_counts[only_staff] += 1
        
        elif weekend:
            # 土日: 1人Y9, 残り9
            candidates = [s for s in staff_working if s not in fixed_shift]
            if not candidates:
                candidates = staff_working
            
            y_candidate = select_y_candidate(candidates, y_counts, prev_day_off)
            for staff in staff_working:
                if staff == y_candidate:
                    result_df.at[idx, staff] = 'Y9'
                    y_counts[staff] += 1
                else:
                    if staff in fixed_shift:
                        result_df.at[idx, staff] = fixed_shift[staff]
                    else:
                        result_df.at[idx, staff] = '9'
        else:
            # 平日: 9時が1人必要、残りは10時
            # 1. 固定シフトとフリーを分類
            fixed_9 = [s for s in staff_working if fixed_shift.get(s) == '9']
            fixed_10 = [s for s in staff_working if fixed_shift.get(s) == '10']
            flexible = [s for s in staff_working if s not in fixed_9 and s not in fixed_10]
            
            final_9 = list(fixed_9)
            final_10 = list(fixed_10)
            
            # 2. 9時担当を確保 (固定9がいなければフリーから選出)
            if not final_9:
                if flexible:
                    picked = random.choice(flexible)
                    final_9.append(picked)
                    flexible.remove(picked)
                # フリーもいなければ9時なし（全員固定10時などの場合）
            
            # 3. 残りのフリーは10時
            final_10.extend(flexible)
            
            # 4. ベース時間を割り当て
            for s in final_9:
                result_df.at[idx, s] = '9'
            for s in final_10:
                result_df.at[idx, s] = '10'
            
            # 5. Y担当決め (固定なしの人を優先)
            all_assigned = final_9 + final_10
            if all_assigned:
                y_candidates = [s for s in all_assigned if s not in fixed_shift]
                if not y_candidates:
                    y_candidates = all_assigned # 全員固定なら仕方なく固定の人から選ぶ
                
                y_candidate = select_y_candidate(y_candidates, y_counts, prev_day_off)
                
                # Yを付加 (時間は維持)
                current_val = result_df.at[idx, y_candidate]
                result_df.at[idx, y_candidate] = 'Y' + current_val
                y_counts[y_candidate] += 1
        
        for staff in staff_columns:
            prev_day_off[staff] = is_off(row[staff])
    
    # 最適化（スワップ）実行
    result_df = optimize_shifts(result_df, staff_columns)
    
    return result_df


def get_y_counts(df: pd.DataFrame, staff_columns: list) -> dict:
    """各スタッフの呼び出し回数を取得"""
    return {
        staff: df[staff].str.contains('Y', na=False).sum()
        for staff in staff_columns
    }


def is_balanced(y_counts: dict) -> bool:
    """呼び出し回数が±1以内かチェック"""
    counts = list(y_counts.values())
    return max(counts) - min(counts) <= 1


def parse_shift_value(value: str) -> tuple:
    """シフト値を（Y有無，時間/休み）に分解する"""
    if pd.isna(value) or value == '':
        return ('', '')
    
    value = str(value)
    
    if value.startswith('Y'):
        return ('Y', value[1:])
    elif value in ['公', '年']:
        return ('', value)
    else:
        return ('', value)


def convert_to_excel_format(df: pd.DataFrame, staff_columns: list) -> pd.DataFrame:
    """結果DataFrameをExcel出力用の形式に変換する"""
    excel_data = []
    
    header = ['', '']
    for staff in staff_columns:
        short_name = staff.replace('スタッフ', '')
        header.extend([short_name, ''])
    
    excel_data.append(header)
    
    for _, row in df.iterrows():
        excel_row = [row['日付'], row['曜日']]
        for staff in staff_columns:
            y_mark, time_value = parse_shift_value(row[staff])
            excel_row.extend([y_mark, time_value])
        excel_data.append(excel_row)
    
    return pd.DataFrame(excel_data)


def generate_shift_schedule(csv_content: str) -> tuple:
    """
    CSVコンテンツから勤務表を生成
    Returns: (excel_bytes, y_counts, attempts)
    """
    df = pd.read_csv(io.StringIO(csv_content), dtype=str)
    staff_columns = [col for col in df.columns if col not in ['日付', '曜日']]
    
    max_attempts = 1000
    best_result = None
    best_diff = float('inf')
    best_y_counts = {}
    attempts = 0
    
    for attempt in range(max_attempts):
        attempts = attempt + 1
        result_df = assign_shifts(df, staff_columns)
        y_counts = get_y_counts(result_df, staff_columns)
        counts = list(y_counts.values())
        diff = max(counts) - min(counts)
        
        if diff < best_diff:
            best_diff = diff
            best_result = result_df
            best_y_counts = y_counts
        
        if is_balanced(y_counts):
            break
    
    # Excel形式に変換
    excel_df = convert_to_excel_format(best_result, staff_columns)
    
    # Excelファイルをメモリに書き込み
    output = io.BytesIO()
    excel_df.to_excel(output, index=False, header=False, engine='openpyxl')
    output.seek(0)
    
    # セル結合処理
    wb = load_workbook(output)
    ws = wb.active
    
    for i, staff in enumerate(staff_columns):
        start_col = 3 + (i * 2)
        end_col = start_col + 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        ws.cell(row=1, column=start_col).alignment = Alignment(horizontal='center')
    
    # 最終的なExcelをメモリに保存
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    
    return final_output, best_y_counts, attempts


# ===== Flaskルート =====

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        password = request.form.get('password', '')
        if password == APP_PASSWORD:
            session['authenticated'] = True
            return redirect(url_for('index'))
        else:
            error = 'パスワードが正しくありません'
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.pop('authenticated', None)
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    return render_template('index.html')


def get_detailed_counts(df: pd.DataFrame, staff_columns: list) -> dict:
    """各スタッフの全シフトタイプの回数を集計"""
    counts = {staff: {'9': 0, '10': 0, '公': 0, '年': 0, 'Y': 0} for staff in staff_columns}
    
    for _, row in df.iterrows():
        for staff in staff_columns:
            val = row[staff]
            if pd.isna(val) or val == '':
                continue
            
            # Yが含まれる場合
            if 'Y' in val:
                counts[staff]['Y'] += 1
                # 時間部分もカウント (Y9 -> 9, Y10 -> 10)
                time_part = val.replace('Y', '')
                if time_part in counts[staff]:
                    counts[staff][time_part] += 1
            else:
                # 通常のシフト記述 (9, 10, 公, 年)
                if val in counts[staff]:
                    counts[staff][val] += 1
                elif val not in ['9', '10', '公', '年']:
                     # 未知の値があればキー追加してカウント（念のため）
                     if val not in counts[staff]:
                         counts[staff][val] = 0
                     counts[staff][val] += 1
                     
    return counts


def count_consecutive_violations(df, staff_columns):
    """連勤違反をカウント (hard>6, soft>5)"""
    hard_count = 0
    soft_count = 0
    for staff in staff_columns:
        streak = 0
        for val in df[staff]:
            if not is_off(val):
                streak += 1
            else:
                if streak > 6: hard_count += (streak - 6)
                elif streak > 5: soft_count += 1
                streak = 0
        if streak > 6: hard_count += (streak - 6)
        elif streak > 5: soft_count += 1
    return hard_count, soft_count


def fix_consecutive_work(df, staff_columns, fixed_mask):
    """6連勤超えを解消する"""
    for staff in staff_columns:
        # Detect streaks > 6
        streaks = [] 
        current_streak = []
        for idx in range(len(df)):
            val = df.at[idx, staff]
            if not is_off(val):
                current_streak.append(idx)
            else:
                if len(current_streak) > 6:
                    streaks.append(current_streak.copy())
                current_streak = []
        if len(current_streak) > 6: streaks.append(current_streak.copy())
        
        for streak in streaks:
            # ブレイクポイントを探す
            breakable = []
            for idx in streak:
                # 固定シフトは除外
                if fixed_mask and fixed_mask.get((idx, staff)):
                    continue
                # 休みを入れても大丈夫な日（人数確保など）
                if can_take_off(df, idx, staff, staff_columns):
                    breakable.append(idx)
            
            if not breakable:
                continue
            
            # 中央付近を優先的にターゲットにする
            import math
            best_idx = breakable[0]
            min_dist = float('inf')
            center = streak[0] + len(streak) / 2
            
            for idx in breakable:
                dist = abs(idx - center)
                if dist < min_dist:
                    min_dist = dist
                    best_idx = idx
            
            target_idx = best_idx
            
            # スワップ先（公休）を探す
            holidays = []
            for idx in range(len(df)):
                if df.at[idx, staff] == '公':
                    holidays.append(idx)
            
            swapped = False
            random.shuffle(holidays)
            for h_idx in holidays:
                # 公休をターゲット日と入れ替え
                # 公休日は基本的に労働可能と仮定（厳密には逆のcan_take_offも要るが省略）
                val_target = df.at[target_idx, staff]
                val_holiday = df.at[h_idx, staff]
                
                df.at[target_idx, staff] = val_holiday
                df.at[h_idx, staff] = val_target
                swapped = True
                break
            
            if not swapped:
                # スワップできない場合は強制的に休みにする
                df.at[target_idx, staff] = '公'

    return df


@app.route('/generate', methods=['POST'])
@login_required
def generate():
    try:
        data = request.json
        year = int(data.get('year', 2026))
        month = int(data.get('month', 1))
        target_holidays = int(data.get('holidayCount', 8))
        staff_names = data.get('staffNames', [])
        schedule = data.get('schedule', [])
        headcount = data.get('headcount', {}) # 人数指定受け取り
        
        if not staff_names:
            return jsonify({'error': 'スタッフ名が入力されていません'}), 400
        
        if not schedule:
            return jsonify({'error': 'スケジュールデータがありません'}), 400
        
        # DataFrameを構築 & マスクを作成
        rows = []
        request_mask = {} # (day_index, staff_name) -> bool
        fixed_mask = {}   # (day_index, staff_name) -> '9' or '10'
        
        for day_idx, day_data in enumerate(schedule):
            row = {
                '日付': day_data['day'],
                '曜日': day_data['dayOfWeek']
            }
            for staff_name in staff_names:
                val = day_data['staff'].get(staff_name, '')
                row[staff_name] = val
                
                # 公 または 年 が入力されている場合は希望休としてマーク
                if val in ['公', '年']:
                    request_mask[(day_idx, staff_name)] = True
                elif val in ['9', '10']:
                    fixed_mask[(day_idx, staff_name)] = val
                    request_mask[(day_idx, staff_name)] = False
                else:
                    request_mask[(day_idx, staff_name)] = False
                    
            rows.append(row)
        
        df = pd.DataFrame(rows)
        
        # 勤務表生成
        staff_columns = staff_names
        
        max_attempts = 1000
        best_result = None
        best_score = (float('inf'), float('inf'), float('inf')) # (hard, soft, diff)
        best_y_counts = {}
        attempts = 0
        
        for attempt in range(max_attempts):
            attempts = attempt + 1
            
            # 公休の自動充填
            df_with_holidays = fill_missing_public_holidays(df.copy(), staff_columns, target_holidays, headcount)
            
            result_df = assign_shifts(df_with_holidays, staff_columns)
            
            # 最適化: 休み明けのY回避
            result_df = optimize_shifts(result_df, staff_columns, fixed_mask)
            
            # 連勤制約の修正
            result_df = fix_consecutive_work(result_df, staff_columns, fixed_mask)
            
            # --- 最終チェック: 固定シフトの強制適用 ---
            for (r_idx, s_name), fix_val in fixed_mask.items():
                current_val = str(result_df.at[r_idx, s_name])
                # 固定値が含まれていなければ強制リセット
                if fix_val not in current_val:
                    result_df.at[r_idx, s_name] = fix_val
            
            y_counts = get_y_counts(result_df, staff_columns)
            counts = list(y_counts.values())
            diff = max(counts) - min(counts)
            
            hard, soft = count_consecutive_violations(result_df, staff_columns)
            current_score = (hard, soft, diff)
            
            if current_score < best_score:
                best_score = current_score
                best_result = result_df.copy()
                best_y_counts = y_counts
            
            # 6連勤超えゼロかつ5連勤超えゼロかつバランス良ければ即終了
            if hard == 0 and soft == 0 and is_balanced(y_counts):
                break
        
        # 詳細集計を取得
        detailed_counts = get_detailed_counts(best_result, staff_columns)

        # Excel形式に変換 & スタイル適用
        # Excel形式に変換 & スタイル適用
        final_output = create_styled_excel(best_result, staff_columns, request_mask)
        
        # セッションに保存
        app.config['LAST_EXCEL'] = final_output
        app.config['LAST_FILENAME'] = f'勤務表{year}{int(month):02d}.xlsx'
        
        # プレビュー用データを作成
        preview_data = []
        for idx, (_, row) in enumerate(best_result.iterrows()):
            row_data = {
                'day': row['日付'],
                'dayOfWeek': row['曜日'],
                'staff': {}
            }
            for staff in staff_columns:
                val = row[staff] if pd.notna(row[staff]) else ''
                is_request = request_mask.get((idx, staff), False)
                row_data['staff'][staff] = {
                    'value': val,
                    'isRequest': is_request
                }
            preview_data.append(row_data)
        
        return jsonify({
            'success': True,
            'attempts': attempts,
            'y_counts': {k: int(v) for k, v in best_y_counts.items()},
            'detailed_counts': detailed_counts,
            'staffNames': staff_columns,
            'preview': preview_data,
            'filename': f'勤務表{year}{int(month):02d}.xlsx'
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'生成エラー: {str(e)}'}), 500


def create_styled_excel(df: pd.DataFrame, staff_columns: list, request_mask: dict = None) -> io.BytesIO:
    """DataFrameからスタイル付きExcelファイルを生成する共通関数"""
    
    # Excel形式に変換
    excel_df = convert_to_excel_format(df, staff_columns)
    
    # Excelファイルをメモリに書き込み
    output = io.BytesIO()
    excel_df.to_excel(output, index=False, header=False, engine='openpyxl')
    output.seek(0)

    # セル結合処理 & レイアウト調整
    wb = load_workbook(output)
    ws = wb.active
    
    # フォント・スタイル設定
    large_font = Font(size=14)
    bold_font = Font(size=14, bold=True)
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # 全体の行高さを設定
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 40
        for cell in row:
            cell.font = large_font
        
    # 列幅の設定
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 6
    
    for i, staff in enumerate(staff_columns):
        start_col_idx = 3 + (i * 2)
        end_col_idx = start_col_idx + 1
        
        y_col_letter = get_column_letter(start_col_idx)     # C, E, G...
        val_col_letter = get_column_letter(end_col_idx)     # D, F, H...
        
        # Yマーク列を狭く
        ws.column_dimensions[y_col_letter].width = 5
        
        # 時間列は少し広めに
        ws.column_dimensions[val_col_letter].width = 10
        
        # ヘッダー結合
        ws.merge_cells(start_row=1, start_column=start_col_idx, end_row=1, end_column=end_col_idx)
        
        # スタイル設定
        # ヘッダー中央揃え
        header_cell = ws.cell(row=1, column=start_col_idx)
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.font = large_font
        
        # データ領域のスタイル
        for row_idx, row in enumerate(range(2, ws.max_row + 1)):
            # row_idx は 0始まり (DataFrameの行インデックスと一致)
            
            # Yマーク列
            cell_y = ws.cell(row=row, column=start_col_idx)
            cell_y.alignment = Alignment(horizontal='center', vertical='center')
            cell_y.font = large_font
            
            # 時間列 / 休み列
            cell_val = ws.cell(row=row, column=end_col_idx)
            cell_val.alignment = Alignment(horizontal='center', vertical='center')
            cell_val.font = large_font
            
            # 希望休ハイライト (request_maskが指定されている場合のみ)
            if request_mask:
                # request_maskのキーは (row_idx, staff_name)
                # キー判定のために staff名が必要
                # staff_columns[i] が現在のスタッフ名
                
                if request_mask.get((row_idx, staff), False):
                     cell_val.font = bold_font
                     cell_val.fill = gray_fill
    
    # A, B列も中央揃え
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')

    
    # 最終的なExcelをメモリに保存
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    
    return final_output


@app.route('/download')
def download():
    # 後方互換性のため残すが、基本はdownload_editedを使う想定
    if 'LAST_EXCEL' not in app.config or app.config['LAST_EXCEL'] is None:
        return jsonify({'error': 'まず勤務表を生成してください'}), 400
    
    excel_bytes = app.config['LAST_EXCEL']
    excel_bytes.seek(0)
    
    filename = app.config.get('LAST_FILENAME', 'shift_schedule.xlsx')
    
    return send_file(
        excel_bytes,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/download_edited', methods=['POST'])
def download_edited():
    data = request.json
    rows = data.get('rows', [])
    staff_names = data.get('staffNames', [])
    filename = data.get('filename', 'worker_schedule.xlsx')
    # memo = data.get('memo', '') # 削除
    
    if not rows or not staff_names:
        return jsonify({'error': 'データが不足しています'}), 400
        
    # DataFrame構築 & request_mask作成
    df_rows = []
    request_mask = {}
    
    for idx, r in enumerate(rows):
        row_dict = {
            '日付': r['day'],
            '曜日': r['dayOfWeek']
        }
        for staff in staff_names:
            val_data = r['staff'].get(staff, '')
            
            # 文字列か辞書かで処理を分ける
            if isinstance(val_data, dict):
                val = val_data.get('value', '')
                is_req = val_data.get('isRequest', False)
                row_dict[staff] = val
                if is_req:
                    request_mask[(idx, staff)] = True
            else:
                row_dict[staff] = val_data
                
        df_rows.append(row_dict)
        
    df = pd.DataFrame(df_rows)
    
    try:
        final_output = create_styled_excel(df, staff_names, request_mask)
        
        return send_file(
            final_output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5002))
    debug = os.environ.get('FLASK_DEBUG', 'true').lower() == 'true'
    app.run(debug=debug, host='0.0.0.0', port=port)
