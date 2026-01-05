#!/usr/bin/env python3

"""
勤務表自動生成ツール
入力CSVから休み情報を読み取り，出勤時間とYマークを自動割り当てする
Author: Taiki Watanabe
"""

import pandas as pd
import random
import sys
from pathlib import Path
# for excel output
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def is_weekend(day_of_week: str) -> bool:
    """土日かどうか判定"""
    return day_of_week in ['土', '日']


def is_off(value: str) -> bool:
    """休みかどうか判定（公，年，または空でない休み表記）"""
    if pd.isna(value) or value == '':
        return False
    return value in ['公', '年']


def get_off_value(value: str) -> str:
    """休みの値を取得（公 or 年）"""
    if pd.isna(value) or value == '':
        return ''
    if value in ['公', '年']:
        return value
    return ''


def assign_shifts(df: pd.DataFrame, staff_columns: list) -> pd.DataFrame:
    """
    勤務表を自動生成する
    
    Args:
        df: 入力DataFrame（日付，曜日，各カメラマンの休み情報）
        staff_columns: カメラマン列名のリスト
    
    Returns:
        出力DataFrame（日付，曜日，各カメラマンの勤務情報）
    """
    result_df = df.copy()
    
    # Yマークのカウント（均等分配用）
    y_counts = {staff: 0 for staff in staff_columns}
    
    # 前日の休み状態を追跡
    prev_day_off = {staff: False for staff in staff_columns}
    
    for idx, row in df.iterrows():
        day_of_week = row['曜日']
        weekend = is_weekend(day_of_week)
        
        # 各カメラマンの休み状態を取得
        staff_off = {}
        staff_working = []
        
        for staff in staff_columns:
            off_value = get_off_value(row[staff])
            if off_value:
                staff_off[staff] = off_value
            else:
                staff_working.append(staff)
        
        # 休みの人はそのまま出力
        for staff, off_value in staff_off.items():
            result_df.at[idx, staff] = off_value
        
        if not staff_working:
            # 全員休みの場合はスキップ
            # 前日の休み状態を更新
            for staff in staff_columns:
                prev_day_off[staff] = is_off(row[staff])
            continue
        
        if len(staff_working) == 1:
            # 1人だけ出勤の場合: 曜日に関係なくY9
            only_staff = staff_working[0]
            result_df.at[idx, only_staff] = 'Y9'
            y_counts[only_staff] += 1
        elif weekend:
            # 土日で2人以上出勤: 1人がY9，他は9時
            y_candidate = select_y_candidate(
                staff_working, y_counts, prev_day_off
            )
            for staff in staff_working:
                if staff == y_candidate:
                    result_df.at[idx, staff] = 'Y9'
                    y_counts[staff] += 1
                else:
                    result_df.at[idx, staff] = '9'
        else:
            # 平日で2人以上出勤: 1人が9時，他が10時，10時の1人にY
            
            # 9時担当をランダムに選ぶ
            nine_am_staff = random.choice(staff_working)
            result_df.at[idx, nine_am_staff] = '9'
            
            # 10時担当（9時以外の出勤者）
            ten_am_staff = [s for s in staff_working if s != nine_am_staff]
            
            # 10時担当の中からYを割り当てる人を選ぶ
            y_candidate = select_y_candidate(
                ten_am_staff, y_counts, prev_day_off
            )
            
            for staff in ten_am_staff:
                if staff == y_candidate:
                    result_df.at[idx, staff] = 'Y10'
                    y_counts[staff] += 1
                else:
                    result_df.at[idx, staff] = '10'
        
        # 前日の休み状態を更新
        for staff in staff_columns:
            prev_day_off[staff] = is_off(row[staff])
    
    return result_df


def select_y_candidate(
    candidates: list, 
    y_counts: dict, 
    prev_day_off: dict
) -> str:
    """
    Yマークを割り当てる候補を選ぶ
    
    優先順位:
    1. 前日が休みでない人
    2. Yの回数が少ない人
    """
    if not candidates:
        raise ValueError("候補者がいません")
    
    if len(candidates) == 1:
        return candidates[0]
    
    # 前日が休みでない人を優先
    not_after_off = [c for c in candidates if not prev_day_off.get(c, False)]
    
    if not_after_off:
        # 休み明けでない人の中でYが少ない人を選ぶ
        target_list = not_after_off
    else:
        # 全員休み明けなら条件を緩和
        target_list = candidates
    
    # Yの回数が最小の人を選ぶ
    min_y = min(y_counts[c] for c in target_list)
    min_y_candidates = [c for c in target_list if y_counts[c] == min_y]
    
    # 同点なら最初の人（安定性のため）
    return min_y_candidates[0]


def get_y_counts(df: pd.DataFrame, staff_columns: list) -> dict:
    """各カメラマンのYマーク回数を取得"""
    return {
        staff: df[staff].str.contains('Y', na=False).sum()
        for staff in staff_columns
    }


def is_balanced(y_counts: dict) -> bool:
    """Yマーク回数が±1以内かチェック"""
    counts = list(y_counts.values())
    return max(counts) - min(counts) <= 1


def parse_shift_value(value: str) -> tuple:
    """
    シフト値を（Y有無，時間/休み）に分解する
    例: 'Y10' -> ('Y', '10'), '9' -> ('', '9'), '公' -> ('', '公')
    """
    if pd.isna(value) or value == '':
        return ('', '')
    
    value = str(value)
    
    if value.startswith('Y'):
        return ('Y', value[1:])  # Y9 -> ('Y', '9'), Y10 -> ('Y', '10')
    elif value in ['公', '年']:
        return ('', value)
    else:
        return ('', value)  # 9 or 10


def convert_to_excel_format(df: pd.DataFrame, staff_columns: list) -> pd.DataFrame:
    """
    結果DataFrameをExcel出力用の形式に変換する
    各カメラマンを2列（左:Y，右:時間）に分割
    """
    # 新しいDataFrameを作成
    excel_data = []
    
    # ヘッダー行を作成
    header = ['', '']  # 日付，曜日
    for staff in staff_columns:
        # カメラマン名から「カメラマン」を除去（あれば）
        short_name = staff.replace('カメラマン', '')
        header.extend([short_name, ''])
    
    excel_data.append(header)
    
    # データ行を作成
    for _, row in df.iterrows():
        excel_row = [row['日付'], row['曜日']]
        for staff in staff_columns:
            y_mark, time_value = parse_shift_value(row[staff])
            excel_row.extend([y_mark, time_value])
        excel_data.append(excel_row)
    
    return pd.DataFrame(excel_data)


def main():
    if len(sys.argv) < 2:
        print("使用方法: python generate_shift.py <入力CSV> [出力ファイル]")
        print("例: python generate_shift.py sample_input.csv output.xlsx")
        print("    python generate_shift.py sample_input.csv output.csv")
        sys.exit(1)
    
    input_file = Path(sys.argv[1])
    
    if len(sys.argv) >= 3:
        output_file = Path(sys.argv[2])
    else:
        # デフォルトはExcel出力
        output_file = input_file.with_name(
            input_file.stem + '_output.xlsx'
        )
    
    # CSV読み込み
    print(f"入力ファイル: {input_file}")
    df = pd.read_csv(input_file, dtype=str)
    
    # カメラマン列を特定（日付，曜日以外の列）
    staff_columns = [col for col in df.columns if col not in ['日付', '曜日']]
    print(f"カメラマン: {', '.join(staff_columns)}")
    
    # 複数試行してYマークが均等になるまで試す
    max_attempts = 1000
    best_result = None
    best_diff = float('inf')
    
    print(f"\n±1以内になるまで最大{max_attempts}回試行します...")
    
    for attempt in range(max_attempts):
        result_df = assign_shifts(df, staff_columns)
        y_counts = get_y_counts(result_df, staff_columns)
        counts = list(y_counts.values())
        diff = max(counts) - min(counts)
        
        if diff < best_diff:
            best_diff = diff
            best_result = result_df
            best_y_counts = y_counts
        
        if is_balanced(y_counts):
            print(f"試行 {attempt + 1}回目で均等な分配を達成")
            break
    else:
        print(f"{max_attempts}回試行しましたが，±1以内にはなりませんでした．")
        print(f"最も均等な結果（差: {best_diff}）を採用します．")
    
    # 出力
    if output_file.suffix.lower() == '.xlsx':
        # Excel出力
        excel_df = convert_to_excel_format(best_result, staff_columns)
        excel_df.to_excel(output_file, index=False, header=False)
        
        wb = load_workbook(output_file)
        ws = wb.active
        
        # 各カメラマンの列を結合（C-D, E-F, G-H, ...）
        for i, staff in enumerate(staff_columns):
            start_col = 3 + (i * 2)  # C=3, E=5, G=7, ...
            end_col = start_col + 1   # D=4, F=6, H=8, ...
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            # 結合セルを中央揃え
            ws.cell(row=1, column=start_col).alignment = Alignment(horizontal='center')
        
        wb.save(output_file)
        print(f"Excel出力: {output_file}")
    else:
        # CSV出力
        best_result.to_csv(output_file, index=False)
        print(f"CSV出力: {output_file}")
    
    # Yマーク集計
    print("\n--- Yマーク集計 ---")
    for staff in staff_columns:
        y_count = best_result[staff].str.contains('Y', na=False).sum()
        print(f"{staff}: {y_count}回")


if __name__ == '__main__':
    main()
